import sqlite3
from datetime import datetime
import os
import logging
import openpyxl
from openpyxl.utils import get_column_letter
import bcrypt
from config import DB_PATH
# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__) # Тоже корневой логгер

# Подключение SQLite
conn = sqlite3.connect(DB_PATH, check_same_thread=False)

# Авторизованные номера и админы (из ENV)
AUTHORIZED_NUMBERS = set(num.strip() for num in (os.getenv("AUTHORIZED_NUMBERS") or "").split(",") if num.strip())
ADMIN_IDS = set(admin.strip() for admin in (os.getenv("ADMINS") or "").split(",") if admin.strip())

# Миграции
def migrate_users_table():
    expected_columns = {
        "telegram_id": "TEXT PRIMARY KEY",
        "number": "TEXT NOT NULL",
        "full_name": "TEXT",
        "role": "TEXT DEFAULT 'user'",
        "auth_time": "TEXT",
        "subscription_active": "INTEGER DEFAULT 0"
    }
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(users)")
    existing = {col[1] for col in cursor.fetchall()}

    for col, definition in expected_columns.items():
        if col not in existing:
            logging.warning(f"[MIGRATION] Добавление колонки users.{col}")
            cursor.execute(f"ALTER TABLE users ADD COLUMN {col} {definition}")

def migrate_documents_table():
    expected_columns = {"title", "description", "revision", "updated_at", "path"}
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM documents LIMIT 1")
        current_cols = {desc[0] for desc in cursor.description}
        missing = expected_columns - current_cols
        if missing:
            logging.warning(f"[MIGRATION] Поля в documents отсутствуют: {missing}")
    except Exception as e:
        logging.error(f"[MIGRATION] Проблема с таблицей documents: {e}")


def remove_user(telegram_id: int):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE telegram_id = ?", (telegram_id,))
    conn.commit()
    conn.close()

def remove_revoked_users(authorized_numbers: set[str]):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT telegram_id, number FROM users")
    users = cur.fetchall()

    removed = []
    for user_id, number in users:
        if number not in authorized_numbers:
            cur.execute("DELETE FROM users WHERE telegram_id = ?", (user_id,))
            removed.append((user_id, number))

    conn.commit()
    conn.close()

    return removed

# Инициализация базы
def init_db():
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            telegram_id TEXT PRIMARY KEY,
            number TEXT NOT NULL,
            full_name TEXT,
            role TEXT DEFAULT 'user',
            auth_time TEXT,
            subscription_active INTEGER DEFAULT 0
        )
    ''')
    migrate_users_table()

    cursor.execute('''
        CREATE VIRTUAL TABLE IF NOT EXISTS documents USING fts5(
            title,
            description,
            revision,
            updated_at,
            path
        )
    ''')
    migrate_documents_table()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS blocked_users (
            user_id INTEGER PRIMARY KEY,
            unblock_time INTEGER
        )
    ''')

    cursor.execute("CREATE INDEX IF NOT EXISTS idx_number ON users (number);")

    conn.commit()
    logging.info("✅ База данных успешно инициализирована.")

# --- Пользователи ---
def add_user(user_id, number, full_name, role="user"):
    hashed_number = bcrypt.hashpw(number.encode(), bcrypt.gensalt()).decode()
    cursor = conn.cursor()
    cursor.execute("""
        REPLACE INTO users (telegram_id, number, full_name, role, auth_time)
        VALUES (?, ?, ?, ?, ?)
    """, (user_id, hashed_number, full_name, role, datetime.utcnow().isoformat()))
    conn.commit()
    logging.info(f"[LOGIN] {user_id=}, {number=}, {full_name=}")

def set_subscription(user_id, active=True):
    cursor = conn.cursor()
    cursor.execute("UPDATE users SET subscription_active = ? WHERE telegram_id = ?", (1 if active else 0, user_id))
    conn.commit()

def get_last_users(limit=10):
    cursor = conn.cursor()
    cursor.execute("""
        SELECT telegram_id, number, full_name, auth_time
        FROM users
        ORDER BY auth_time DESC
        LIMIT ?
    """, (limit,))
    return cursor.fetchall()

# --- Блокировки ---
def add_block(user_id, unblock_time):
    cursor = conn.cursor()
    cursor.execute("INSERT OR REPLACE INTO blocked_users (user_id, unblock_time) VALUES (?, ?)", (user_id, unblock_time))
    conn.commit()

def remove_block(user_id):
    cursor = conn.cursor()
    cursor.execute("DELETE FROM blocked_users WHERE user_id = ?", (user_id,))
    conn.commit()

def is_blocked(user_id):
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM blocked_users WHERE user_id = ?", (user_id,))
    return cursor.fetchone() is not None

def remove_expired_blocks():
    now = int(datetime.now().timestamp())
    cursor = conn.cursor()
    cursor.execute("DELETE FROM blocked_users WHERE unblock_time <= ?", (now,))
    conn.commit()

def clear_blocks():
    cursor = conn.cursor()
    cursor.execute("DELETE FROM blocked_users")
    conn.commit()

# --- Авторизация ---
def is_authorized(user_id):
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM users WHERE telegram_id = ?", (user_id,))
    return cursor.fetchone() is not None

def get_user_role(user_id):
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE telegram_id = ?", (user_id,))
    result = cursor.fetchone()
    return result[0] if result else None

def is_number_taken(number):
    cursor = conn.cursor()
    cursor.execute("SELECT number FROM users")
    for (hashed_number,) in cursor.fetchall():
        if bcrypt.checkpw(number.encode(), hashed_number.encode()):
            return True
    return False

def is_same_user(user_id, number):
    cursor = conn.cursor()
    cursor.execute("SELECT telegram_id, number FROM users")
    for row_user_id, hashed_number in cursor.fetchall():
        if bcrypt.checkpw(number.encode(), hashed_number.encode()) and str(row_user_id) == str(user_id):
            return True
    return False

def get_user_number(user_id: int) -> str | None:
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute("SELECT number FROM users WHERE telegram_id = ?", (user_id,))
        row = cur.fetchone()
    return row[0] if row else None

def export_users_to_excel(filename="users_export.xlsx"):
    cursor = conn.cursor()
    cursor.execute("""
        SELECT telegram_id, number, full_name, role, auth_time, subscription_active
        FROM users
        ORDER BY auth_time DESC
    """)
    users = cursor.fetchall()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    headers = ["Telegram ID", "Number", "Full Name", "Role", "Auth Time", "Subscribed"]
    sheet.append(headers)

    for row in users:
        sheet.append(list(row))

    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    workbook.save(filename)
    return filename
